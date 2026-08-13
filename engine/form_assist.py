"""
공문 양식 채우기 — 통합 엔드포인트.

1. 여러 파일 수신 → 모두 텍스트 추출
2. 출력 양식 파일 지정 → 빈칸 추출
3. LLM에게 전체 맥락 + 빈칸 + 교사 지시 전달
4. 양식에 값 주입 → 완성 파일 반환
"""
from __future__ import annotations

import json
import os
import re
import tempfile
from pathlib import Path
from typing import Optional


def run_form_assist(
    files: list[dict],           # [{"path": str, "name": str}, ...]
    instruction: str,            # 교사 지시사항
    output_file_idx: int = -1,   # 출력 양식 파일 인덱스 (-1이면 자동 감지)
    page_range: str = "",        # "2-3" 등 (빈칸이면 전체)
    llm_provider: str = "auto",
    llm_model: str = "",         # "provider/model" 형식 또는 모델명
    llm_config: dict = None,
    live_mode: bool = True,      # True면 한/글 COM으로 실시간 채우기
    progress_cb=None,
    log_cb=None,
    output_dir: str = "",        # 출력 파일 저장 경로 (빈값 = 바탕화면)
    hwp_elements: list[dict] = None,  # InitScan 스캔 결과 (셀 ID 기반 채우기용)
) -> dict:
    """통합 공문 양식 채우기."""

    def log(msg):
        if log_cb:
            log_cb(msg)

    def progress(val):
        if progress_cb:
            progress_cb(val)

    if not files:
        raise ValueError("파일이 없습니다.")

    temp_dir = tempfile.mkdtemp(prefix="tf_form_")

    # ── 1. 모든 파일 텍스트 추출 ──
    progress(0.1)
    log("파일 텍스트 추출 중...")

    extracted = []  # [{"name", "ext", "text", "path", "is_form"}]
    form_exts = {".xlsx", ".xls", ".hwpx", ".hwp"}

    for i, f in enumerate(files):
        fpath = f["path"]
        fname = f.get("name", Path(fpath).name)
        ext = Path(fpath).suffix.lower()

        text = _extract_text(fpath, ext, temp_dir, log)
        is_form = ext in form_exts

        extracted.append({
            "index": i,
            "name": fname,
            "ext": ext,
            "text": text,
            "path": fpath,
            "is_form": is_form,
        })
        log(f"  {fname} ({ext}) → {len(text)}자")

    # ── 2. 출력 양식 결정 ──
    progress(0.3)

    output_template = None
    if output_file_idx >= 0 and output_file_idx < len(extracted):
        output_template = extracted[output_file_idx]
    else:
        # 자동 감지: 양식 확장자 파일 중 마지막 것
        for e in reversed(extracted):
            if e["is_form"]:
                output_template = e
                break

    # ── 3. 양식 유형 판별 + 빈칸/구조 준비 ──
    #   fill_mode: none | excel | hwpx_grid | hwp_com | hwp_text
    #   - hwpx_grid = 병합-인지 그리드(COM-free, 벤치 495/495 검증 경로) + 셀ID enum 강제
    fill_mode = "none"
    json_schema: Optional[dict] = None
    blanks_json = ""
    grid_doc = None
    grid_fields: list[dict] = []
    blank_ids: set = set()

    if output_template:
        t_ext = output_template["ext"]
        if t_ext in (".xlsx", ".xls"):
            fill_mode = "excel"
            log(f"양식 빈칸 추출: {output_template['name']}")
            try:
                from nodes.form_extract.main import execute as extract_fn
                ctx = {"temp_dir": temp_dir, "progress": lambda x: None, "log": log}
                ex = extract_fn(
                    inputs={"파일": output_template["path"]},
                    params={"include_filled": False},
                    context=ctx,
                )
                blanks_json = ex.get("빈칸목록", "")
                blanks = json.loads(blanks_json) if blanks_json else []
                log(f"  빈칸 {len(blanks)}개 감지")
            except Exception as e:
                log(f"  빈칸 추출 실패: {e}")
                blanks_json = "[]"
        elif t_ext == ".hwpx":
            # 병합-인지 그리드 빈칸(행헤더×열헤더 라벨) + 본문 밑줄 블랭크 + 누름틀
            # — 전부 COM 불필요.
            try:
                from engine.hwpml.hwpx_grid import (
                    parse_hwpx, extract_blank_fields, extract_body_blanks,
                )
                grid_doc = parse_hwpx(output_template["path"])
                grid_fields = [
                    f for f in extract_blank_fields(grid_doc, include_filled=True)
                    if f.get("value_type") in ("text", "paren", "colon")
                    and _is_fillable(f) and not _is_signature_field(f)
                ]
                grid_fields += extract_body_blanks(output_template["path"])  # 본문 ___
                grid_fields += _extract_hwpx_fields(output_template["path"])  # 누름틀
                if grid_fields:
                    fill_mode = "hwpx_grid"
                    blank_ids = {f["id"] for f in grid_fields}
                    json_schema = _build_fill_schema(sorted(blank_ids))
                    log(f"양식 그리드: 표 {len(grid_doc.tables)}개, 채울 빈칸 {len(grid_fields)}개")
                else:
                    fill_mode = "hwp_text"
                    log("표/본문/누름틀 빈칸 없음 — 텍스트 치환 경로")
            except Exception as e:
                log(f"HWPX 그리드 분석 실패 → 텍스트 경로: {e}")
                fill_mode = "hwp_text"
        elif t_ext == ".hwp":
            # 레거시 바이너리: COM InitScan 결과(hwp_elements)로 셀ID 기반 채우기.
            if hwp_elements:
                fill_mode = "hwp_com"
                json_schema = _build_fill_schema([str(e["id"]) for e in hwp_elements])
                log(f"HWP 구조 스캔: {len(hwp_elements)}개 요소")
            else:
                fill_mode = "hwp_text"

    # ── 4. LLM 프롬프트 구성 ──
    progress(0.5)
    log("AI에게 전달 중...")

    # 맥락 텍스트 (양식 제외한 모든 파일)
    context_parts = []
    for e in extracted:
        if e is not output_template and e["text"]:
            context_parts.append(f"### {e['name']}\n{e['text']}")
    context_text = "\n\n---\n\n".join(context_parts) if context_parts else "(참고 문서 없음)"

    prompt = f"""당신은 교사의 공문 양식을 채우는 비서입니다.

## 참고 문서
{context_text}

## 교사 지시사항
{instruction if instruction else "(없음)"}
"""
    range_note = f"\n### 작성 범위: {page_range}\n" if page_range else ""

    if fill_mode == "excel":
        prompt += f"""
## 출력 양식: {output_template['name']}
### 빈칸 목록
{blanks_json}
{range_note}
위 참고 문서와 교사 지시를 바탕으로 각 빈칸에 적절한 값을 채우세요.
반드시 JSON으로만 답하세요: {{"셀참조": "값", ...}}
빈칸 목록의 cell_ref를 키로 사용하세요. 설명 없이 JSON만 반환.
"""
    # hwpx_grid는 표 단위 청킹(_plan_grid_fill)으로 별도 처리 — 프롬프트 불필요
    elif fill_mode == "hwp_com":
        cell_desc = _format_hwp_elements(hwp_elements)
        prompt += f"""
## 출력 양식: {output_template['name']}
### 문서 구조 (셀 ID + 현재 내용)
{cell_desc}
{range_note}
위 참고 문서와 교사 지시를 바탕으로, 빈칸이 있는 셀을 채우세요.
- 빈칸(○○○, ___, 공란, 빈 값, 미입력)이거나 채워야 할 셀만 포함하세요.
- 이미 올바른 값이 들어있는 셀은 건드리지 마세요.
- id는 위 구조의 id 값(숫자 문자열)을 그대로 사용하세요.
"""
    elif fill_mode == "hwp_text":
        prompt += f"""
## 출력 양식: {output_template['name']}
### 양식 텍스트
{output_template['text'][:5000]}
{range_note}
위 참고 문서와 교사 지시를 바탕으로, 이 양식의 빈칸을 채우세요.
반드시 JSON으로만 답하세요: {{"찾을텍스트": "바꿀텍스트", ...}}
- 양식에서 빈칸(○○○, ___, 공란, 예시 텍스트 등)을 찾아 실제 값으로 바꾸세요.
- 표의 빈 셀이나 미완성 내용도 포함하세요. 설명 없이 JSON만 반환하세요.
"""
    else:
        prompt += """
출력 양식이 지정되지 않았습니다.
참고 문서와 교사 지시를 바탕으로 요청된 내용을 작성하세요.
마크다운 형식으로 답하세요.
"""

    # ── 5. LLM 호출 (양식이면 json_schema 강제; hwpx_grid는 청킹 계획) ──
    progress(0.6)
    if fill_mode == "hwpx_grid":
        grid_fill_data = _plan_grid_fill(grid_doc, grid_fields, context_text, instruction,
                                         page_range, llm_provider, llm_model,
                                         llm_config or {}, log)
        llm_response = ""  # 청크별 내부 처리 (단일 응답 없음)
    else:
        llm_response = _call_llm(prompt, llm_provider, llm_model, llm_config or {},
                                 json_schema=json_schema)
        log(f"AI 응답: {len(llm_response)}자")

    # ── 6. 저장 경로 + 모드별 주입 ──
    progress(0.8)
    result = {"text": llm_response, "file": None}
    save_dir = _resolve_save_dir(output_dir, temp_dir)

    if fill_mode == "hwpx_grid":
        try:
            from engine.hwpml.hwpx_grid import (
                ID_RE, BODY_ID_RE, fill_hwpx_cells, relocate_below_markers,
            )
            fill_data = grid_fill_data
            log(f"그리드 채우기: {len(fill_data)}개 항목")
            if fill_data:
                # 셀ID(그리드)/본문블랭크ID/누름틀명(정확일치) 3-way 분리 — 누름틀은
                # 라벨 퍼지매칭이 아닌 정확 이름 일치로만 채워 무관 셀 과충전을 막는다.
                grid_map = {k: v for k, v in fill_data.items() if ID_RE.match(k)}
                body_map = {k: v for k, v in fill_data.items() if BODY_ID_RE.match(k)}
                field_map = {k: v for k, v in fill_data.items()
                             if not ID_RE.match(k) and not BODY_ID_RE.match(k)}
                # '이하빈칸' 마커 자동 이동 (데이터가 마커 클리어보다 우선)
                extra = relocate_below_markers(grid_doc, grid_map, log=log)
                merged = {**extra, **grid_map}
                out_path = os.path.join(save_dir, Path(output_template["name"]).stem + "_완성.hwpx")
                filled = fill_hwpx_cells(output_template["path"], out_path, merged,
                                         log=log, field_map=field_map or None,
                                         body_map=body_map or None)
                if filled:
                    result["file"] = out_path
                    log(f"완성 파일: {out_path}")

                    # ── 재추출 검증 + 미반영 1회 재시도 (설계 §1) ──
                    verified, missing = _verify_hwpx_fill(out_path, fill_data)
                    if missing:
                        log(f"검증: {len(verified)}/{len(fill_data)} 반영, {len(missing)}개 미반영 — 재시도")
                        retry = _retry_fill(missing, grid_fields, context_text, instruction,
                                            llm_provider, llm_model, llm_config or {}, log)
                        if retry:
                            r_grid = {k: v for k, v in retry.items() if ID_RE.match(k)}
                            r_body = {k: v for k, v in retry.items() if BODY_ID_RE.match(k)}
                            r_field = {k: v for k, v in retry.items()
                                       if not ID_RE.match(k) and not BODY_ID_RE.match(k)}
                            fill_hwpx_cells(out_path, out_path, r_grid, log=log,
                                            field_map=r_field or None, body_map=r_body or None)
                            verified, missing = _verify_hwpx_fill(out_path, fill_data)
                    result["verified"] = len(verified)
                    result["missing"] = list(missing)
                    result["text"] = (f"{len(verified)}개 빈칸을 채웠습니다."
                                      + (f" ({len(missing)}개는 확인이 필요합니다)" if missing else ""))
                    log(f"검증 완료: {len(verified)}/{len(fill_data)} 반영"
                        + (f", 미반영 {len(missing)}개: {list(missing)[:5]}" if missing else ""))
                else:
                    log("주입된 셀 없음 — 셀ID 매칭 실패")
            else:
                log("LLM이 채울 항목을 반환하지 않았습니다")
        except Exception as e:
            import traceback
            log(f"그리드 채우기 실패: {e}")
            log(f"  {traceback.format_exc().splitlines()[-1]}")

    elif fill_mode == "hwp_com":
        # COM 전용 스레드 필요 → 데이터만 반환 (서버 라우트가 fill_hwp_by_cells 실행)
        fill_data = _parse_fill_response(llm_response, {str(e["id"]) for e in hwp_elements})
        if fill_data:
            result["fill_data"] = fill_data
            result["template_path"] = output_template["path"]
            result["save_dir"] = save_dir
            log(f"HWP 채우기 데이터 {len(fill_data)}개 항목 준비 (COM)")
        else:
            log("LLM이 채울 항목을 반환하지 않았습니다")

    elif fill_mode == "hwp_text" and output_template["ext"] == ".hwp":
        # .hwp 텍스트 폴백은 form_fill._fill_hwp(win32com PutFieldText)로 가는데
        # ① 전용 COM 스레드(deps._com_pool)를 벗어나 실행되고(행/충돌 위험)
        # ② find/replace 키가 누름틀 필드명과 안 맞아 조용히 미충전된다.
        # → .hwp는 자동 채우기 대신 안내 텍스트만 반환(정상 채움은 hwp_com 경로 담당).
        log("HWP 구조 스캔 결과 없음 — 자동 채우기 생략(텍스트만 반환)")

    elif fill_mode in ("excel", "hwp_text"):
        try:
            fill_data = _extract_json_from_response(llm_response)
            if fill_data is None:
                raise json.JSONDecodeError("JSON not found", llm_response, 0)
            log(f"양식 주입: {len(fill_data)}개 항목")
            from nodes.form_fill.main import execute as fill_fn
            output_name = Path(output_template["name"]).stem + "_완성"
            fill_result = fill_fn(
                inputs={
                    "양식파일": output_template["path"],
                    "채울내용": json.dumps(fill_data, ensure_ascii=False),
                },
                params={"output_name": output_name},
                context={"temp_dir": save_dir, "progress": lambda x: None, "log": log},
            )
            result["file"] = fill_result.get("파일")
            if result["file"]:
                log(f"완성 파일: {result['file']}")
        except json.JSONDecodeError:
            log("JSON 파싱 실패 — AI가 JSON 형식으로 답하지 않았습니다")
            log(f"  응답 앞 200자: {llm_response[:200]}")
        except Exception as e:
            import traceback
            log(f"양식 주입 실패: {e}")
            log(f"  {traceback.format_exc().splitlines()[-1]}")

    progress(1.0)
    return result


# ── 그리드 배치 계획 (라이브 기록용 — run_form_assist와 동일 프롬프트/스키마) ──

def plan_hwpx_grid_fill(
    form_path: str,
    instruction: str,
    context_text: str = "",
    llm_provider: str = "local",
    llm_model: str = "",
    llm_config: dict = None,
    log=None,
) -> dict:
    """hwpx 양식의 채움 계획을 gemma가 결정 — {셀ID|본문ID|누름틀명: 값} 반환.

    run_form_assist의 hwpx_grid 모드와 같은 라벨그리드+json_schema(enum) 방식.
    COM 불필요(파일 파싱만) — 라이브 기록(grid_live)과 파일 채움 어느 쪽에도 사용 가능.
    """
    def _log(msg):
        if log:
            log(msg)

    from engine.hwpml.hwpx_grid import (
        parse_hwpx, extract_blank_fields, extract_body_blanks,
    )

    grid_doc = parse_hwpx(form_path)
    grid_fields = [
        f for f in extract_blank_fields(grid_doc, include_filled=True)
        if f.get("value_type") in ("text", "paren", "colon")
        and _is_fillable(f)
    ]
    grid_fields += extract_body_blanks(form_path)
    grid_fields += _extract_hwpx_fields(form_path)
    if not grid_fields:
        _log("채울 빈칸 없음 (표/본문/누름틀)")
        return {"fill_data": {}, "grid_doc": grid_doc, "blank_ids": set()}

    blank_ids = {f["id"] for f in grid_fields}
    _log(f"양식 그리드: 표 {len(grid_doc.tables)}개, 채울 빈칸 {len(grid_fields)}개")

    fill_data = _plan_grid_fill(grid_doc, grid_fields, context_text, instruction,
                                "", llm_provider, llm_model, llm_config or {}, _log)
    labels = {f["id"]: f.get("label", "") for f in grid_fields}
    return {"fill_data": fill_data, "grid_doc": grid_doc, "blank_ids": blank_ids,
            "labels": labels}


# ── 라이브 HWP 채우기 ──

def _is_windows():
    import platform
    return platform.system() == "Windows"

def _connect_hwp(form_path: str, log) -> "tuple[any, bool]":
    """한/글 인스턴스에 연결. 이미 열린 문서면 활성화만. (COM 스레드 전용)"""
    import pythoncom
    pythoncom.CoInitialize()
    from pyhwpx import Hwp

    hwp = Hwp(visible=True)
    target_name = Path(form_path).name
    abs_path = os.path.abspath(form_path)

    need_open = True
    try:
        xdocs = hwp.XHwpDocuments
        for i in range(xdocs.Count):
            doc = xdocs.Item(i)
            full = getattr(doc, 'FullName', '') or ''
            # 정확 일치만 — 부분문자열 매칭('양식.hwpx' ⊂ '제출용 양식.hwpx')은
            # 엉뚱한 문서를 활성화해 그 창에 기록하는 사고로 이어진다.
            same_path = os.path.normcase(os.path.abspath(full)) == os.path.normcase(abs_path)
            same_name = Path(full).name.lower() == target_name.lower()
            if full and (same_path or same_name):
                doc.SetActive_XHwpDocument()
                need_open = False
                log(f"이미 열린 문서에서 작업: {target_name}")
                break
    except Exception:
        pass

    if need_open:
        import time
        log(f"한/글에서 파일 열기: {target_name}")
        hwp.Open(abs_path)
        time.sleep(0.5)

    return hwp, need_open


def scan_hwp_structure(form_path: str, log, timeout: float = 15.0) -> list[dict]:
    """한/글 문서를 InitScan으로 스캔하여 셀/문단 구조 반환. (COM 스레드 전용)

    Inline AI 역공학 기반: init_scan → get_text → get_pos 패턴.
    list_id > 0 이면 표 셀 내부.
    """
    import pythoncom
    pythoncom.CoInitialize()

    hwp, _ = _connect_hwp(form_path, log)

    elements = []
    block_id = 0

    log("문서 구조 스캔 (InitScan)...")
    try:
        import time as _time
        hwp.init_scan(option=4, range=0x0077)

        prev_pos = None
        scan_start = _time.monotonic()
        for _ in range(20000):  # 무한루프 방지
            if _time.monotonic() - scan_start > timeout:
                log(f"스캔 타임아웃 ({timeout:.0f}초)")
                break
            state, text = hwp.get_text()
            if state <= 1:
                break

            hwp.move_pos(201)
            pos = hwp.get_pos()

            clean = (text or "").replace("\r\n", "").replace("\r", "")
            if prev_pos == pos and not clean.strip():
                continue
            prev_pos = pos

            list_id, para_id, char_pos = pos
            elem_type = "td" if list_id > 0 else "text"

            elements.append({
                "id": str(block_id),
                "type": elem_type,
                "text": clean,
                "pos": [list_id, para_id, char_pos],
                "list_id": list_id,
            })
            block_id += 1

        hwp.release_scan()
        log(f"  스캔 완료: {len(elements)}개 요소 ({sum(1 for e in elements if e['type']=='td')}개 표 셀)")
    except Exception as e:
        log(f"  스캔 실패: {e}")

    return elements


def fill_hwp_by_cells(form_path: str, fill_data: dict, elements: list[dict],
                      log, output_dir: str = "") -> str:
    """셀 ID 기반 커서 이동 → SelectAll → insert_text 로 채우기. (COM 스레드 전용)

    Inline AI 역공학 기반: set_pos → is_cell → SelectAll → insert_text 패턴.
    """
    import pythoncom
    pythoncom.CoInitialize()
    import time

    try:
        hwp, _ = _connect_hwp(form_path, log)

        # 위치 맵: id → [list, para, pos]
        pos_map = {str(e["id"]): e for e in elements}

        filled = 0
        log(f"채우기 시작: {len(fill_data)}개 항목")

        for cell_id, new_text in fill_data.items():
            val = str(new_text).strip() if not isinstance(new_text, dict) else str(new_text.get("value", "")).strip()
            if not val:
                continue

            elem = pos_map.get(str(cell_id))
            if not elem:
                log(f"  #{cell_id}: 위치 없음, 건너뜀")
                continue

            list_id, para_id, char_pos = elem["pos"]

            try:
                hwp.set_pos(list_id, para_id, char_pos)

                if list_id > 0 and hwp.is_cell():
                    # 표 셀: 전체 선택 후 삽입
                    hwp.SelectAll()
                    hwp.insert_text(val)
                else:
                    # 본문 문단: 시작~끝 선택 후 삽입
                    hwp.MoveParaBegin()
                    hwp.MoveSelParaEnd()
                    hwp.insert_text(val)

                filled += 1
                log(f"  셀 #{cell_id} → {val[:40]}")
                time.sleep(0.03)
            except Exception as e:
                log(f"  셀 #{cell_id} 실패: {e}")

        log(f"한/글 {filled}개 항목 채우기 완료")

        # 다른 이름으로 저장
        output_name = Path(form_path).stem + "_완성"
        save_dir = output_dir or ""
        if not save_dir:
            desktop = Path.home() / "Desktop"
            if desktop.exists():
                save_dir = str(desktop)
            else:
                for p in Path.home().glob("OneDrive*/바탕*화면"):
                    if p.is_dir():
                        save_dir = str(p)
                        break
        if not save_dir:
            save_dir = os.path.dirname(form_path)

        ext = Path(form_path).suffix.lower()
        output_path = os.path.join(save_dir, f"{output_name}{ext}")
        hwp.SaveAs(os.path.abspath(output_path))
        log(f"저장: {output_path}")

        return output_path

    except Exception as e:
        log(f"pyhwpx 채우기 실패: {e}")
        import traceback
        log(f"  {traceback.format_exc().splitlines()[-1]}")
        return ""


# ── HWP 요소 포매팅 ──

def _format_hwp_elements(elements: list[dict]) -> str:
    """스캔된 HWP 요소를 LLM 프롬프트용 텍스트로 변환."""
    lines = []
    for e in elements:
        eid = e["id"]
        etype = "표셀" if e["type"] == "td" else "본문"
        text = e.get("text", "").strip()
        label = f"(빈칸)" if not text else text[:200]
        lines.append(f'<cell id="{eid}" type="{etype}">{label}</cell>')
    return "\n".join(lines)


# ── 그리드 빈칸 채우기 헬퍼 ──

# 명백한 "빈칸 표기" 글자만 (동그라미·밑줄류). 체크박스(□■▢)·대시·마침표·단일
# 글자는 선택 상태/실데이터일 수 있어 제외 — 실데이터 덮어쓰기 사고 방지.
_PLACEHOLDER_CHARS = set("○◯〇_＿")
# 셀 전체가 이 단어일 때만 자리표시자로 인정(부분문자열 매칭은 '…기입하였음' 같은
# 실데이터를 오인하므로 금지).
_PLACEHOLDER_WORDS = {"예시", "미기재", "미입력", "기입란", "작성란", "기입", "작성"}


def _is_placeholder(text: str) -> bool:
    """'채워 넣으라'는 명백한 자리표시자만 인정 (예: ○○○, ___, '예시', '미기재').

    보수적으로 판정 — 값이 있는 셀을 빈칸으로 오인해 덮어쓰는 사고를 막는다.
    (빈 셀은 is_empty로 이미 잡히므로, 여기서 놓쳐도 실제 빈칸 누락은 없다.)
    """
    s = (text or "").strip()
    if not s or len(s) > 12:
        return False
    if s in _PLACEHOLDER_WORDS:  # 전체 일치만 (부분문자열 오인 방지)
        return True
    core = s.replace(" ", "")
    # 자리표시 글자로만 구성 + 길이 2+ (단일 ○/O 같은 OX 데이터 오인 방지)
    return len(core) >= 2 and all(ch in _PLACEHOLDER_CHARS for ch in core)


def _is_fillable(field: dict) -> bool:
    """빈 셀은 채움 대상. 값이 있으면 명백한(짧은) 자리표시자일 때만.

    부분 슬롯(괄호형/콜론말미형)은 추출기가 이미 판정한 입력칸 —
    채움도 원문 보존 삽입(_compose_slot_value)이라 덮어쓰기 위험 없음.
    """
    if field.get("is_empty"):
        return True
    if field.get("value_type") in ("paren", "colon"):
        return True
    return _is_placeholder(field.get("current_value", ""))


def _extract_hwpx_fields(path: str) -> list[dict]:
    """HWPX 누름틀(form-field) 목록 → 채움 대상 필드.

    id = 필드명(=form_fill 레거시 경로의 매칭 키). 표 셀 그리드와 별개로 채운다.
    """
    import zipfile
    from lxml import etree

    fields: list[dict] = []
    seen: set = set()
    try:
        with zipfile.ZipFile(path, "r") as zf:
            secs = sorted(n for n in zf.namelist()
                          if "section" in n.lower() and n.endswith(".xml"))
            for sec in secs:
                try:
                    root = etree.fromstring(zf.read(sec))
                except etree.XMLSyntaxError:
                    continue
                for elem in root.iter():
                    tag = (etree.QName(elem.tag).localname
                           if "}" in str(elem.tag) else str(elem.tag))
                    if tag in ("fieldBegin", "FIELDBEGIN"):
                        name = elem.get("name", "") or elem.get("Name", "")
                        if not name or name in seen:
                            continue
                        seen.add(name)
                        fields.append({
                            "id": name, "label": f"[누름틀] {name}",
                            "current_value": "", "is_empty": True,
                            "value_type": "field",
                        })
    except Exception:
        pass
    return fields


def _build_fill_schema(ids) -> dict:
    """{채움:[{id∈enum, 값:str}]} 강제 스키마 — 로컬은 GBNF가 셀ID를 못 틀리게 한다.

    maxItems=빈칸 수: 소형모델이 같은 셀ID를 반복하며 루프해 max_tokens를
    소진하고 **응답 전체를 잃는** 실측(§41c: 33빈칸 문서에서 절단 → 0필드).
    배열 길이를 빈칸 수로 묶으면 루프가 문법 차원에서 끝난다.
    """
    ids = list(ids)
    return {
        "type": "object",
        "properties": {
            "채움": {
                "type": "array",
                "maxItems": max(1, len(ids)),
                "items": {
                    "type": "object",
                    "properties": {
                        "id": {"type": "string", "enum": list(ids)},
                        "값": {"type": "string"},
                    },
                    "required": ["id", "값"],
                },
            }
        },
        "required": ["채움"],
    }


def _render_blank_list(fields: list[dict]) -> str:
    """빈칸 목록을 'id : 라벨 (현재값)' 줄로 렌더 (LLM이 의미 매칭할 대상)."""
    lines = []
    for f in fields:
        cur = (f.get("current_value") or "").strip()
        note = f"  (현재: {cur[:20]})" if cur else ""
        lines.append(f"- {f['id']} : {f.get('label', '')}{note}")
    return "\n".join(lines)


def _parse_fill_response(text: str, valid_ids=None) -> dict:
    """LLM 응답 → {셀ID: 값} 딕셔너리.

    스키마 강제형({채움:[{id,값}]}), 배열형, 평면 dict({id:값}) 모두 흡수.
    valid_ids가 주어지면 그 집합 밖의 셀ID는 버린다(소프트 강제/환각 방어).
    """
    valid = set(valid_ids) if valid_ids else None
    out: dict = {}
    obj = _extract_json_from_response(text)

    items = []
    if isinstance(obj, dict) and isinstance(obj.get("채움"), list):
        items = obj["채움"]
    elif isinstance(obj, list):
        items = obj
    elif isinstance(obj, dict):
        # 평면 {셀ID: 값} (API 소프트 강제 또는 구형 응답)
        for k, v in obj.items():
            if isinstance(v, dict):
                v = v.get("값", v.get("value", ""))
            if v is not None:
                out[str(k).strip()] = str(v)

    for it in items:
        if not isinstance(it, dict):
            continue
        cid = str(it.get("id") or it.get("셀ID") or it.get("cell_id") or "").strip()
        if not cid:
            continue
        val = it.get("값", it.get("value", it.get("text", "")))
        # 중복 id는 **첫 값 채택**(§41c 실측: 소형모델이 같은 셀ID를 반복 출력하며
        # 루프 — 19항목 중 고유 12개, 한 ID 7회. 마지막-우선이면 루프 쓰레기가
        # 정답을 덮어쓴다).
        if cid in out:
            continue
        out[cid] = "" if val is None else str(val)

    if valid is not None:
        out = {k: v for k, v in out.items() if k in valid}
    return out


# 청크 하나의 그리드 렌더 문자 예산 (소형 모델 ctx 8192 안전 여유)
def _norm_label(s: str) -> str:
    """라벨 정규화 — 공백 제거('학 생 명'→'학생명')."""
    return re.sub(r"\s+", "", str(s or ""))


def _parse_instruction_pairs(instruction: str) -> dict:
    """지시문에서 명시적 라벨→값 쌍 추출 (Stage1 사전 매칭 §42).

    지원: ①"라벨은(는) 값, 라벨은(는) 값." 나열형(FormAssist 정형 지시)
         ②"라벨: 값" 줄 단위 콜론형(자유 지시).
    오추출 쌍은 빈칸 라벨과 매칭 실패로 무해하다(내장 안전핀) —
    실제 채움은 라벨이 유일하게 일치할 때만 일어난다.
    """
    pairs: dict = {}
    text = instruction or ""

    # ① 은(는) 나열형 — 마커로 자르고, 값|다음라벨은 마지막 ", "에서 분리
    #    ("2,480,000원"처럼 공백 없는 콤마는 값에 안전하게 남는다)
    segs = re.split(r"(?:은|는)\((?:는|은)\)\s*", text)
    if len(segs) >= 2:
        def _tail_label(s: str) -> str:
            t = re.split(r"[.\n]", s)[-1]
            if ", " in t:
                t = t.rsplit(", ", 1)[-1]
            return t.strip()

        label = _tail_label(segs[0])
        for i in range(1, len(segs)):
            seg = segs[i]
            if i < len(segs) - 1:
                value, _, nxt = seg.rpartition(", ")
                if not value:
                    value, nxt = seg, ""
            else:
                # 문장 종결 "." 1개 제거 (값 자체가 "26."으로 끝나는 날짜는
                # 종결과 구분 불가한 극소수 케이스 — 수용)
                value, nxt = re.sub(r"\.\s*$", "", seg), ""
            if label and value.strip():
                pairs.setdefault(_norm_label(label), value.strip())
            label = _tail_label(nxt) if nxt else ""

    # ② 콜론형 — "라벨: 값. 라벨: 값." (한 줄 다중 쌍 + 줄 단위 혼용).
    #   콜론(연속 "::")을 분리자로, 라벨 = 직전 조각의 문장 꼬리,
    #   값 = 다음 라벨 직전까지. 날짜값("2026. 3. 28.")의 내부 ". "는
    #   마지막 분리자 rsplit이라 보존된다.
    # 시간("10:00")·비율 등 숫자:숫자 콜론은 분리자가 아님
    cms = list(re.finditer(r"[:：]+(?=\s)|(?<!\d)[:：]+(?!\d)", text))
    if cms:
        def _tail2(s: str) -> str:
            t = re.split(r"\.\s+|\n", s)[-1]
            if ", " in t:
                t = t.rsplit(", ", 1)[-1]
            return t.strip(" -·•※\t")

        for j, m in enumerate(cms):
            lab = _tail2(text[(cms[j - 1].end() if j else 0):m.start()])
            nxt = cms[j + 1].start() if j + 1 < len(cms) else len(text)
            seg = text[m.end():nxt]
            if "\n" in seg and j + 1 < len(cms):
                val = seg.split("\n", 1)[0]
            elif j + 1 < len(cms):
                val = seg.rsplit(". ", 1)[0] if ". " in seg else seg
            else:
                val = re.sub(r"\.\s*$", "", seg)
            lab, val = lab.strip(), val.strip()
            if lab and val and len(lab) <= 60 and not val.startswith("//"):
                pairs.setdefault(_norm_label(lab), val)
    return pairs


_SIGNATURE_RE = re.compile(r"서\s*명|\(인\)|（인）|날인|사인|signature", re.I)


def _is_signature_field(f: dict) -> bool:
    """서명·날인 칸은 자필 영역 — 채움 후보에서 원천 제외(§42k, Opus급 갭 실측).

    Opus는 라벨 의미로 알아서 비웠지만 소형모델은 채우려 든다 →
    판단을 코드로 옮긴다. (지시문이 서명 값을 명시하면 사전 매칭이
    라벨 유일 일치로 여전히 채울 수 있음 — 후보 제외는 LLM 경로만.)
    """
    return bool(_SIGNATURE_RE.search(str(f.get("label") or "")))


def _scan_freeform_pairs(instruction: str, fields: list) -> dict:
    """자유문장 라벨-스캔 — 문서 라벨을 지시문에서 찾아 조사(에/엔/은/는/을/를,
    콜론) 뒤의 값을 다음 라벨/구두점 경계까지 취한다. 정형 파서 0쌍일 때 전용."""
    text = instruction or ""
    occs = []
    for f in fields:
        lab = f.get("label") or ""
        nl = _norm_label(lab)
        if len(nl) < 3:
            continue
        pat = r"\s*".join(re.escape(c) for c in lab if not c.isspace())
        for m in re.finditer(pat, text):
            occs.append((m.start(), m.end(), nl))
    if not occs:
        return {}
    occs.sort(key=lambda x: (x[0], -(x[1] - x[0])))
    sel, last = [], -1
    for s, e, nl in occs:
        if s >= last:
            sel.append((s, e, nl))
            last = e
    pairs = {}
    for j, (s, e, nl) in enumerate(sel):
        rest = text[e:]
        m = re.match(r"\s*(?:에는|에도|엔|에|은|는|을|를|도)?\s*[:：=]?\s*", rest)
        vs = e + (m.end() if m else 0)
        ve = sel[j + 1][0] if j + 1 < len(sel) else len(text)
        val = text[vs:ve]
        # 값 경계: 콤마·마침표·개행 앞까지 (날짜 "2026. 7. 8." 내부 ". "은 보존
        # 못 하므로 콤마·개행 우선, 마침표는 뒤에 공백+한글이 올 때만 경계)
        val = re.split(r"[,\n]|(?<=[.\d])\.\s+(?=[가-힣])", val)[0]
        val = re.sub(r"\s*(?:로|으로)?\s*(?:넣어|넣고|기입|입력|채워|적어|해)\S*\s*$", "", val).strip(" .")
        if val and len(val) <= 60:
            pairs.setdefault(nl, val)
    return pairs


def prematch_fields(instruction: str, fields: list,
                    all_fields: list | None = None) -> tuple:
    """지시문 명시 쌍을 코드가 100% 정밀로 선채움 — "지능을 코드로" (§42 Stage1).

    1차 = 정규화 완전일치 + 후보 유일, 2차 = 포함 일치 + 후보 유일.
    중복 라벨(어느 칸인지 원리상 모호)·다중 후보는 LLM에 넘긴다.
    반환: (선채움 {id:값}, 잔여 필드 목록).
    """
    pairs = _parse_instruction_pairs(instruction)
    # (실험 §42i: 자유문장 라벨-스캔 폴백은 5/5 오배치로 기각 — 자유문장은
    # 라벨을 바꿔 말하므로 리터럴 매칭 불가. 의미 매칭은 LLM 몫으로 유지.)
    if not pairs:
        return {}, fields
    by_norm: dict = {}
    for f in fields:
        by_norm.setdefault(_norm_label(f.get("label")), []).append(f)

    # 연결 오염 가드(§42b 오배치 23건의 근원): 값 안에 다른 빈칸 라벨이
    # 통째로 들어있으면 "값+다음라벨" 접합 쓰레기 — 기각(LLM 폴백).
    # 라벨 지식은 (주어지면) 문서 전체 빈칸에서 — 대상 밖 라벨 접합도 탐지
    label_src = all_fields if all_fields else fields
    known = list({_norm_label(f.get("label")) for f in label_src
                  if len(_norm_label(f.get("label"))) >= 4})

    def _tainted(nv: str, own: str) -> bool:
        return any(nl != own and nl in nv for nl in known)

    # 값을 라벨 경계에서 재절단: 값 꼬리가 알려진 라벨이면 잘라 회수
    def _rescue(val: str) -> str:
        nv = _norm_label(val)
        for nl in sorted(known, key=len, reverse=True):
            if nv.endswith(nl) and len(nv) > len(nl):
                pat = r"\s*".join(re.escape(c) for c in nl)
                m = re.search(pat + r"\s*$", val)
                if m and m.start() > 0:
                    return val[:m.start()].rstrip().rstrip(",.").rstrip()
        return val

    cleaned = {}
    for nl, val in pairs.items():
        val = _rescue(val)
        if not _tainted(_norm_label(val), nl):
            cleaned[nl] = val
    pairs = cleaned

    filled: dict = {}
    matched_labels = set()
    for nl, val in pairs.items():
        cand = by_norm.get(nl, [])
        if nl and len(cand) == 1:
            filled[cand[0]["id"]] = val
            matched_labels.add(nl)
    for nl, val in pairs.items():
        if nl in matched_labels or len(nl) < 2:
            continue
        cands = [f for f in fields
                 if f["id"] not in filled and len(_norm_label(f.get("label"))) >= 2
                 and (nl in _norm_label(f.get("label"))
                      or _norm_label(f.get("label")) in nl)]
        if len(cands) == 1:
            filled[cands[0]["id"]] = val
    residual = [f for f in fields if f["id"] not in filled]
    return filled, residual


_GRID_CHUNK_CHARS = 7000
# 청크 하나의 빈칸 수 상한 — 문자 예산과 별개. 소형 모델은 한 응답에 담을
# 항목이 많으면 중도 포기하거나 같은 셀ID를 반복한다(§41c 실측).
_GRID_CHUNK_FIELDS = 15


def _plan_grid_fill(grid_doc, grid_fields: list, context_text: str, instruction: str,
                    page_range: str, provider: str, model: str, config: dict, log) -> dict:
    """그리드 빈칸 채움 계획 — 표 단위 청킹으로 대형 양식의 렌더 절단을 방지.

    작은 양식은 1청크(기존 동작과 동일). 표별 렌더+빈칸을 문자 예산 내로 묶어
    청크마다 그 청크의 셀ID enum만 강제 → 수백 빈칸도 절단 없이 전부 노출.
    본문 밑줄·누름틀은 라벨 목록만으로 마지막 청크에 싣는다(렌더 불필요).
    반환: 병합된 {id:값}.
    """
    from engine.hwpml.hwpx_grid import ID_RE

    # ── 0. 결정적 사전 매칭 — 지시문의 명시 쌍은 코드가 확정(오배치 0) ──
    pre_fill: dict = {}
    if config.get("fill_prematch", True):
        pre_fill, grid_fields = prematch_fields(instruction, grid_fields)
        if pre_fill:
            log(f"사전 매칭: {len(pre_fill)}개 코드 확정, LLM 잔여 {len(grid_fields)}개")
        if not grid_fields:
            log(f"배치 결정: {len(pre_fill)}개 항목 (전부 사전 매칭)")
            return pre_fill

    # 필드를 표별 / 기타(본문·누름틀)로 분류
    by_table: dict = {}
    misc: list = []
    for f in grid_fields:
        m = ID_RE.match(str(f["id"]))
        if m and f.get("value_type") in ("text", "paren", "colon"):
            by_table.setdefault(f"s{m.group(1)}_t{m.group(2)}", []).append(f)
        else:
            misc.append(f)

    # 섹션 = (렌더, 필드들). 표는 개별 render(), 기타는 렌더 없이 목록만.
    # 빈칸이 많은 표는 **같은 렌더를 공유하며 필드만 분할**한다 — 문자 예산만으로는
    # 안 걸리는 밀집 양식(§41c 실측: 3.5천자에 빈칸 33·40개)에서 모델이 중도
    # 포기해 10/33·6/40만 채우던 문제. 렌더는 반복해도 프롬프트 캐시가 흡수한다.
    sections: list = []
    for grid in grid_doc.tables:
        fs = by_table.get(grid.key)
        if not fs:
            continue
        render = grid.render(mark_blanks=True)
        for i in range(0, len(fs), _GRID_CHUNK_FIELDS):
            part = fs[i:i + _GRID_CHUNK_FIELDS]
            # v13: 거대 표는 청크별 행-윈도 렌더 — 타깃 행 주변만 노출해
            # 절단 없이 전 행 커버(학습 v13과 동일 규칙 = 분포 일치).
            if len(render) > 3000:
                rows = set()
                for f in part:
                    m2 = ID_RE.match(str(f["id"]))
                    if m2:
                        rows.add(int(m2.group(3)))
                sections.append((grid.render_row_window(rows, ctx=1), part))
            else:
                sections.append((render, part))
    for i in range(0, len(misc), _GRID_CHUNK_FIELDS):
        sections.append(("", misc[i:i + _GRID_CHUNK_FIELDS]))

    # 문자 예산 + 필드 수 예산으로 청크 패킹 (한 섹션이 초과해도 최소 1섹션은 담는다)
    chunks: list = []
    cur_r, cur_f, cur_len = [], [], 0
    for render, fs in sections:
        over = (cur_len + len(render) > _GRID_CHUNK_CHARS
                or len(cur_f) + len(fs) > _GRID_CHUNK_FIELDS)
        if cur_f and over:
            chunks.append((cur_r, cur_f))
            cur_r, cur_f, cur_len = [], [], 0
        if render not in cur_r:          # 같은 표를 쪼갠 경우 렌더 중복 방지
            cur_r.append(render)
            cur_len += len(render)
        cur_f.extend(fs)
    if cur_f:
        chunks.append((cur_r, cur_f))

    if len(chunks) > 1:
        log(f"대형 양식 — 표 단위 {len(chunks)}청크로 분할 채움")

    range_note = f"\n### 작성 범위: {page_range}\n" if page_range else ""
    fill_data: dict = {}
    for i, (renders, fields) in enumerate(chunks):
        ids = {f["id"] for f in fields}
        schema = _build_fill_schema(sorted(ids))
        grid_render = "\n".join(r for r in renders if r)
        if len(grid_render) > _GRID_CHUNK_CHARS:
            grid_render = grid_render[:_GRID_CHUNK_CHARS] + "\n…(생략)"
        struct = f"### 문서 표 구조 (빈칸은 {{셀ID}} 로 표시됨)\n{grid_render}\n\n" if grid_render else ""
        prompt = f"""당신은 교사의 공문 양식을 채우는 비서입니다.

## 참고 문서
{context_text if context_text else "(참고 문서 없음)"}

## 교사 지시사항
{instruction if instruction else "(없음)"}
{range_note}
{struct}### 채워야 할 빈칸 ({len(fields)}개)
{_render_blank_list(fields)}

위 참고 문서와 교사 지시를 바탕으로, 각 빈칸에 알맞은 값을 넣으세요.
- 빈칸 라벨(행 이름 × 열 이름)의 의미에 맞는 값을 채우세요.
- 값을 알 수 없거나 채울 필요가 없는 빈칸은 생략하세요.
- 이미 의미 있는 값이 들어 있는 칸은 그대로 두세요(비어 있을 때만 채움).
- 값에는 빈칸에 들어갈 내용만 쓰세요 — 라벨이나 기호("(인)", "성명:" 등)를 반복하지 마세요.
- 서명·날인 칸과 표 제목 칸은 채우지 마세요. 지시문이 비우라는 칸은 생략하세요.
- id는 위 '채워야 할 빈칸' 목록의 id를 정확히 그대로 쓰세요.
"""
        if len(chunks) > 1:
            log(f"  청크 {i+1}/{len(chunks)}: 빈칸 {len(fields)}개")
        votes = int(config.get("fill_votes", 1) or 1)
        if votes <= 1:
            resp = _call_llm(prompt, provider, model, config, json_schema=schema,
                             fill=True)
            fill_data.update(_parse_fill_response(resp, ids))
        else:
            # k-표결 — temp>0 샘플 k개에서 셀별 다수결(동수는 최초 응답 우선).
            # 소형모델의 간헐 오독을 상호 상쇄 (§42 Stage1).
            from collections import Counter
            vote_cfg = dict(config)
            vote_cfg["temperature"] = float(config.get("vote_temperature", 0.6))
            per_id: dict = {}
            for _ in range(votes):
                r = _call_llm(prompt, provider, model, vote_cfg,
                              json_schema=schema, fill=True)
                for k, v in _parse_fill_response(r, ids).items():
                    per_id.setdefault(k, Counter())[str(v)] += 1
            fill_data.update(
                {k: c.most_common(1)[0][0] for k, c in per_id.items()})
    # 라벨 에코·예시 복사 기각(§42h 오류 해부: 회수 가능 11/19건) —
    # 예측값이 자기 라벨과 같거나(에코), 라벨 안에 통째로 들어있는 긴 문자열
    # (라벨의 "(예시)…" 복사)이면 무효. 짧은 정당값("남"/"여" 등)은 보호.
    labels = {f["id"]: _norm_label(f.get("label")) for f in grid_fields}
    def _suspect(cid, val):
        nv, nl = _norm_label(val), labels.get(cid, "")
        if not nv or not nl:
            return False
        if nv == nl:
            return True
        if len(nv) >= 6 and nv in nl:
            return True
        # 역방향: 값 안에 자기 라벨이 통째로("12×담당교사 서명" 류 변형 포함)
        return len(nl) >= 4 and nl in nv
    dropped = [k for k, v in fill_data.items() if _suspect(k, v)]
    for k in dropped:
        del fill_data[k]
    if dropped:
        log(f"라벨 에코/예시 기각: {len(dropped)}건 (검토 패널로 위임)")
    fill_data = {**fill_data, **pre_fill}  # 사전 매칭이 항상 우선
    log(f"배치 결정: {len(fill_data)}개 항목 (사전 매칭 {len(pre_fill)}개 포함)")
    return fill_data


def _verify_hwpx_fill(out_path: str, fill_data: dict) -> tuple[dict, dict]:
    """완성본을 재파싱해 각 값이 실제 반영됐는지 확인 (설계 §1 '재추출 검증').

    반환: (verified {id:값}, missing {id:값}). 빈 값('')은 검증 대상 제외.
    """
    from engine.hwpml.hwpx_grid import parse_hwpx, ID_RE, BODY_ID_RE

    def _n(s: str) -> str:
        return "".join((s or "").split())

    doc = parse_hwpx(out_path)
    cells = {f"{g.key}_r{r}_c{c}": cell.text
             for g in doc.tables for (r, c), cell in g.cells.items()}
    flow = _n(doc.render_text())

    verified, missing = {}, {}
    for k, v in fill_data.items():
        val = str(v).strip()
        if not val:
            continue
        nv = _n(val)
        if ID_RE.match(str(k)):
            ok = nv in _n(cells.get(k, ""))
        elif BODY_ID_RE.match(str(k)):
            ok = nv in flow
        else:  # 누름틀명
            ok = nv in flow or any(nv in _n(t) for t in cells.values())
        (verified if ok else missing)[k] = v
    return verified, missing


def _retry_fill(missing: dict, grid_fields: list, context_text: str, instruction: str,
                provider: str, model: str, config: dict, log) -> dict:
    """미반영 빈칸만 LLM에 다시 요청 (1회 재시도). 반환: {id:값} (유효 id 필터)."""
    fields = [f for f in grid_fields if f["id"] in missing]
    if not fields:
        return {}
    ids = {f["id"] for f in fields}
    schema = _build_fill_schema(sorted(ids))
    prompt = f"""당신은 교사의 공문 양식을 채우는 비서입니다.

## 참고 문서
{context_text if context_text else "(참고 문서 없음)"}

## 교사 지시사항
{instruction if instruction else "(없음)"}

## 아직 못 채운 빈칸 ({len(fields)}개) — 아래만 정확히 다시 채우세요
{_render_blank_list(fields)}

- 각 빈칸 라벨의 의미에 맞는 값만 쓰세요. 값에는 라벨·기호를 반복하지 마세요.
- id는 위 목록의 id를 정확히 그대로 쓰세요.
"""
    resp = _call_llm(prompt, provider, model, config, json_schema=schema)
    return _parse_fill_response(resp, ids)


def _resolve_save_dir(output_dir: str, temp_dir: str) -> str:
    """저장 경로: 설정값 > 바탕화면 > OneDrive 바탕화면 > temp."""
    if output_dir:
        return output_dir
    desktop = Path.home() / "Desktop"
    if desktop.exists():
        return str(desktop)
    for p in Path.home().glob("OneDrive*/바탕*화면"):
        if p.is_dir():
            return str(p)
    return temp_dir


# ── JSON 추출 (LLM 응답에서) ──

def _extract_json_from_response(text: str) -> dict | None:
    """LLM 응답에서 JSON dict를 추출. 여러 방법으로 시도."""
    import re

    # 1. ```json ... ``` 블록
    m = re.search(r"```(?:json)?\s*\n?(.*?)```", text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1).strip())
        except json.JSONDecodeError:
            pass

    # 2. 전체가 JSON
    try:
        return json.loads(text.strip())
    except json.JSONDecodeError:
        pass

    # 3. { ... } 블록 찾기 (가장 바깥 중괄호)
    brace_start = text.find("{")
    brace_end = text.rfind("}")
    if brace_start != -1 and brace_end > brace_start:
        try:
            return json.loads(text[brace_start:brace_end + 1])
        except json.JSONDecodeError:
            pass

    return None


# ── 텍스트 추출 ──

def _extract_text(path: str, ext: str, temp_dir: str, log) -> str:
    """파일에서 텍스트 추출."""
    try:
        if ext == ".pdf":
            import pymupdf
            doc = pymupdf.open(path)
            return "\n\n".join(page.get_text() for page in doc)

        if ext in (".hwpx",):
            from nodes.hwpx_to_md.main import execute as hwpx_fn
            r = hwpx_fn({"파일": path}, {}, {"temp_dir": temp_dir, "progress": lambda x: None, "log": log})
            return r.get("텍스트", "")

        if ext in (".docx",):
            from nodes.docx_to_md.main import execute as docx_fn
            r = docx_fn({"파일": path}, {}, {"temp_dir": temp_dir, "progress": lambda x: None, "log": log})
            return r.get("텍스트", "")

        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                parts = []
                for ws in wb:
                    parts.append(f"=== 시트: {ws.title} ===")
                    for row in ws.iter_rows(max_row=min(ws.max_row or 0, 200), values_only=False):
                        cells = [str(c.value) for c in row if c.value is not None]
                        if cells:
                            parts.append(" | ".join(cells))
                wb.close()
                return "\n".join(parts)
            except (IndexError, Exception):
                # openpyxl 스타일 파싱 실패 → pandas fallback
                log(f"openpyxl 실패, pandas로 텍스트 추출")
                try:
                    import pandas as pd
                    dfs = pd.read_excel(path, sheet_name=None, header=None)
                    parts = []
                    for name, df in dfs.items():
                        parts.append(f"=== 시트: {name} ===")
                        for _, row in df.iterrows():
                            cells = [str(v) for v in row if pd.notna(v)]
                            if cells:
                                parts.append(" | ".join(cells))
                    return "\n".join(parts)
                except Exception as e2:
                    log(f"pandas 추출도 실패: {e2}")
                    return f"[엑셀 텍스트 추출 실패: {Path(path).name}]"

        if ext in (".txt", ".md", ".csv"):
            return Path(path).read_text(encoding="utf-8", errors="ignore")

        # ODT, PPT 등 — 간단한 텍스트 추출 시도
        if ext == ".odt":
            import zipfile
            from lxml import etree
            with zipfile.ZipFile(path) as zf:
                content = zf.read("content.xml")
                root = etree.fromstring(content)
                texts = root.itertext()
                return "\n".join(t.strip() for t in texts if t.strip())

        if ext in (".pptx",):
            from nodes.pptx_to_md.main import execute as pptx_fn
            r = pptx_fn({"파일": path}, {}, {"temp_dir": temp_dir, "progress": lambda x: None, "log": log})
            return r.get("텍스트", "")

        if ext == ".hwp":
            # HWP 바이너리 — olefile/pyhwp로 텍스트 추출 (COM 불필요)
            try:
                import olefile
                ole = olefile.OleFileIO(path)
                if ole.exists("PrvText"):
                    data = ole.openstream("PrvText").read()
                    return data.decode("utf-16-le", errors="ignore")
                # BodyText 스트림에서 추출 시도
                texts = []
                for stream in ole.listdir():
                    name = "/".join(stream)
                    if "BodyText" in name or "PrvText" in name:
                        raw = ole.openstream(stream).read()
                        text = raw.decode("utf-16-le", errors="ignore")
                        # 제어 문자 제거
                        text = "".join(c for c in text if c.isprintable() or c in "\n\r\t")
                        if text.strip():
                            texts.append(text.strip())
                ole.close()
                if texts:
                    return "\n\n".join(texts)
            except Exception as e2:
                log(f"  olefile 추출 실패: {e2}")
            # fallback: pyhwp
            try:
                from hwp5.hwp5txt import extract_text
                import io
                buf = io.StringIO()
                extract_text(path, buf)
                return buf.getvalue()
            except Exception:
                pass
            return f"[HWP 텍스트 추출 실패: {Path(path).name}]"

        return f"[지원하지 않는 형식: {ext}]"

    except Exception as e:
        log(f"  텍스트 추출 실패 ({Path(path).name}): {e}")
        return f"[추출 실패: {Path(path).name}]"


# ── LLM 호출 ──

def _call_llm(prompt: str, provider: str, model: str, config: dict,
              json_schema: dict | None = None, fill: bool = False) -> str:
    """llm_manager를 통한 멀티 프로바이더 LLM 호출.

    json_schema가 있으면 로컬은 GBNF 강제, API는 소프트 강제(generate_chat 내부).
    """
    from engine import deps

    mgr = deps.llm_manager
    if not mgr:
        raise RuntimeError("LLM Manager가 초기화되지 않았습니다.")

    # model 파라미터에서 provider/model 분리 (예: "openai/gpt-4.1")
    if model and "/" in model:
        provider, model_name = model.split("/", 1)
    elif model:
        model_name = model
    else:
        model_name = ""

    # provider가 auto이면 llm_manager에게 위임
    if provider == "auto":
        provider = mgr._pick_provider()

    # 프로바이더별 기본 모델
    if not model_name:
        defaults = {
            "openai": "gpt-4.1",
            "claude": "claude-sonnet-4-6",
            "gemini": "gemini-2.5-flash",
        }
        model_name = defaults.get(provider, "gpt-4.1")

    messages = [
        {"role": "system", "content": "공문 양식을 채우는 비서입니다. 지시에 정확히 따릅니다."},
        {"role": "user", "content": prompt},
    ]

    # 채움 요청 표시(로컬 전용): fill_lora 프리로드 시 채움 어댑터(id1)만 켬 §42f
    if fill:
        mgr._fill_request = True
    try:
        return mgr.generate_chat(
            messages,
            max_tokens=config.get("max_tokens", 4000),
            temperature=config.get("temperature", 0.2),
            provider=provider,
            model=model_name,
            json_schema=json_schema,
        )
    finally:
        mgr._fill_request = False
