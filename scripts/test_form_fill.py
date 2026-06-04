"""
공문 양식 채우기 e2e 테스트.

1. 샘플 Excel 양식 생성 (출장신청서)
2. 샘플 공문 본문 텍스트 작성
3. form_extract로 빈칸 추출
4. LLM으로 빈칸 내용 생성
5. form_fill로 원본에 주입
6. 결과 확인
"""
import json
import os
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

# ═══════════════════════════════════════════
#  1. 샘플 양식 생성
# ═══════════════════════════════════════════

print("=== 1. 샘플 양식 생성 ===")

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "출장신청서"

# 스타일
header_font = Font(bold=True, size=14)
label_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
gray_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

# 제목
ws.merge_cells("A1:F1")
ws["A1"] = "출 장 신 청 서"
ws["A1"].font = Font(bold=True, size=18)
ws["A1"].alignment = Alignment(horizontal="center")

# 양식 필드 (라벨 + 빈칸)
fields = [
    (3, "A", "성명", "B"),
    (3, "C", "직위", "D"),
    (3, "E", "소속", "F"),
    (4, "A", "출장지", "B"),
    (4, "C", "출장 기간", "D"),
    (4, "E", "교통편", "F"),
    (5, "A", "출장 목적", "B"),
    (6, "A", "세부 일정", "B"),
    (7, "A", "예상 경비", "B"),
    (7, "C", "경비 항목", "D"),
    (8, "A", "비고", "B"),
    (10, "A", "위와 같이 출장을 신청합니다.", None),
    (12, "A", "신청일", "B"),
    (12, "C", "신청자", "D"),
    (12, "E", "결재", "F"),
]

for row, label_col, label, value_col in fields:
    cell = ws[f"{label_col}{row}"]
    cell.value = label
    cell.font = label_font
    cell.fill = gray_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if value_col:
        vc = ws[f"{value_col}{row}"]
        vc.value = ""  # 빈칸
        vc.border = thin_border

# 열 너비
for col in ["A", "C", "E"]:
    ws.column_dimensions[col].width = 14
for col in ["B", "D", "F"]:
    ws.column_dimensions[col].width = 22

# 병합 (출장 목적, 세부 일정은 넓게)
ws.merge_cells("B5:F5")
ws.merge_cells("B6:F6")
ws.merge_cells("B8:F8")

sample_form = os.path.join(os.path.dirname(__file__), "..", "data", "sample_출장신청서.xlsx")
wb.save(sample_form)
print(f"  양식 저장: {sample_form}")

# ═══════════════════════════════════════════
#  2. 공문 본문 (맥락)
# ═══════════════════════════════════════════

print("\n=== 2. 공문 본문 ===")

context = """
[경기도교육청 공문 제2026-1234호]

제목: 2026학년도 1학기 수학과 교과연구회 워크숍 참가 안내

1. 경기도교육청 중등교육과-1234(2026.4.8.)호와 관련입니다.
2. 2026학년도 1학기 수학과 교과연구회 워크숍을 아래와 같이 실시하오니,
   참가를 희망하는 교사는 첨부 양식(출장신청서)을 작성하여 4월 15일까지 제출하여 주시기 바랍니다.

가. 일시: 2026년 5월 10일(금) 09:00 ~ 17:00
나. 장소: 경기도교육연수원 (수원시 장안구 경수대로 1079)
다. 대상: 도내 고등학교 수학 교사
라. 내용: AI 활용 수학 교육 방법론, 데이터 기반 학습 분석
마. 경비: 교통비 실비 지급 (대중교통 기준)
바. 교통: 자가용 또는 대중교통 이용

신청자 정보:
- 성명: 신병철
- 직위: 교사
- 소속: 경기과학고등학교
"""

print(f"  공문 텍스트: {len(context)}자")

# ═══════════════════════════════════════════
#  3. form_extract — 빈칸 추출
# ═══════════════════════════════════════════

print("\n=== 3. 빈칸 추출 ===")

from nodes.form_extract.main import execute as extract_execute

temp_dir = tempfile.mkdtemp(prefix="tf_test_")
ctx = {
    "temp_dir": temp_dir,
    "progress": lambda x: None,
    "log": lambda msg: print(f"  [extract] {msg}"),
}

result = extract_execute(
    inputs={"파일": sample_form},
    params={"include_filled": False},
    context=ctx,
)

fields_json = result["빈칸목록"]
form_text = result["원본텍스트"]
fields_list = json.loads(fields_json)

print(f"  빈칸 {len(fields_list)}개 추출됨")
for f in fields_list[:5]:
    print(f"    {f['cell_ref']}: {f['label']}")
if len(fields_list) > 5:
    print(f"    ... (총 {len(fields_list)}개)")

# ═══════════════════════════════════════════
#  4. LLM으로 빈칸 채우기
# ═══════════════════════════════════════════

print("\n=== 4. LLM으로 빈칸 생성 ===")

secrets = json.load(open(os.path.join(os.path.dirname(__file__), "..", "data", ".secrets.json"), encoding="utf-8"))
api_key = secrets.get("openai_api_key", "")

from openai import OpenAI
client = OpenAI(api_key=api_key)

fill_prompt = f"""당신은 교사의 공문 양식을 채우는 비서입니다.

## 공문 본문 (맥락)
{context}

## 양식의 빈칸 목록
{fields_json}

## 양식 전체 텍스트
{form_text}

## 지시
위 공문 내용을 바탕으로 각 빈칸에 적절한 값을 채우세요.
신청자 정보는 공문에 나와있습니다.

반드시 아래 JSON 형식으로만 답하세요. 설명 없이 JSON만:
{{"셀참조": "값", ...}}

예시: {{"B3": "신병철", "D3": "교사", ...}}
"""

response = client.chat.completions.create(
    model="gpt-4.1",
    messages=[
        {"role": "system", "content": "공문 양식을 채우는 비서입니다. JSON으로만 답합니다."},
        {"role": "user", "content": fill_prompt},
    ],
    max_tokens=2000,
    temperature=0.2,
)

fill_text = response.choices[0].message.content
# JSON 추출
if "```json" in fill_text:
    fill_text = fill_text.split("```json")[1].split("```")[0]
elif "```" in fill_text:
    fill_text = fill_text.split("```")[1].split("```")[0]

fill_data = json.loads(fill_text)
print(f"  LLM 생성: {len(fill_data)}개 항목")
for k, v in fill_data.items():
    print(f"    {k}: {v}")

# ═══════════════════════════════════════════
#  5. form_fill — 값 주입
# ═══════════════════════════════════════════

print("\n=== 5. 양식에 값 주입 ===")

from nodes.form_fill.main import execute as fill_execute

result = fill_execute(
    inputs={"양식파일": sample_form, "채울내용": json.dumps(fill_data, ensure_ascii=False)},
    params={"output_name": "완성_출장신청서"},
    context={
        "temp_dir": temp_dir,
        "progress": lambda x: None,
        "log": lambda msg: print(f"  [fill] {msg}"),
    },
)

output_file = result["파일"]
print(f"  완성 파일: {output_file}")

# 결과를 data/ 폴더에 복사
import shutil
final_path = os.path.join(os.path.dirname(__file__), "..", "data", "완성_출장신청서.xlsx")
shutil.copy2(output_file, final_path)
print(f"  최종 저장: {final_path}")

# ═══════════════════════════════════════════
#  6. 결과 검증
# ═══════════════════════════════════════════

print("\n=== 6. 결과 검증 ===")

wb2 = openpyxl.load_workbook(final_path)
ws2 = wb2.active
print(f"  시트: {ws2.title}")
for row in range(3, 13):
    for col in ["A", "B", "C", "D", "E", "F"]:
        v = ws2[f"{col}{row}"].value
        if v:
            print(f"    {col}{row} = {v}")

print("\n=== 완료 ===")
