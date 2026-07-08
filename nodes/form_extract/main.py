"""
м–‘мӢқ л№Ҳм№ё м¶”м¶ң л…ёл“ң.

XLSX: openpyxlлЎң л№Ҳ м…Җ + мқём ‘ лқјлІЁ м¶”м¶ң
HWPX: ZIPвҶ’XML нҢҢмӢұ, л№Ҳ н…ҚмҠӨнҠё н•„л“ң нғҗмғү
HWP: н•ң/кёҖ COM API (Windows м „мҡ©)
"""

import json
import os
import zipfile
from pathlib import Path


def execute(inputs: dict, params: dict, context: dict) -> dict:
    file_path = inputs["нҢҢмқј"]
    if not Path(file_path).exists():
        raise FileNotFoundError(f"нҢҢмқј м—ҶмқҢ: {file_path}")

    include_filled = params.get("include_filled", False)
    ext = Path(file_path).suffix.lower()

    # кө¬нҳ• .xls(OLE л°”мқҙл„ҲлҰ¬)лҠ” openpyxl лҜём§Җмӣҗ вҶ’ нҳјлһҖмҠӨлҹ¬мҡҙ BadZipFile лҢҖмӢ  лӘ…нҷ•нһҲ м•ҲлӮҙ
    if ext == ".xls":
        raise ValueError(
            "кө¬нҳ• .xls нҳ•мӢқмқҖ м§Җмӣҗн•ҳм§Җ м•ҠмҠөлӢҲлӢӨ. н•ң/кёҖмқҙлӮҳ м—‘м…Җм—җм„ң .xlsxлЎң м ҖмһҘн•ң л’Ө мӮ¬мҡ©н•ҳм„ёмҡ”."
        )

    context["progress"](0.1)

    if ext == ".xlsx":
        try:
            fields, full_text = _extract_xlsx(file_path, include_filled, context)
        except (IndexError, Exception) as e:
            context["log"](f"openpyxl нҢҢмӢұ мӢӨнҢЁ, lxml м§Ғм ‘ нҢҢмӢұмңјлЎң мһ¬мӢңлҸ„: {e}")
            fields, full_text = _extract_xlsx_raw(file_path, include_filled, context)
    elif ext == ".hwpx":
        fields, full_text = _extract_hwpx(file_path, include_filled, context)
    elif ext == ".hwp":
        fields, full_text = _extract_hwp(file_path, include_filled, context)
    else:
        raise ValueError(f"м§Җмӣҗн•ҳм§Җ м•ҠлҠ” нҳ•мӢқ: {ext}")

    context["progress"](0.9)
    context["log"](f"л№Ҳм№ё {len(fields)}к°ң м¶”м¶ң мҷ„лЈҢ")
    context["progress"](1.0)

    return {
        "л№Ҳм№ёлӘ©лЎқ": json.dumps(fields, ensure_ascii=False, indent=2),
        "мӣҗліён…ҚмҠӨнҠё": full_text,
    }


# в”Җв”Җ XLSX в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ

def _extract_xlsx(path: str, include_filled: bool, context) -> tuple[list, str]:
    import openpyxl

    context["log"]("м—‘м…Җ м–‘мӢқ л¶„м„қ мӨ‘...")
    wb = openpyxl.load_workbook(path, data_only=True)
    fields = []
    text_parts = []
    field_id = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"=== мӢңнҠё: {sheet_name} ===")

        # лӘЁл“  м…Җ мҠӨмә”
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            row_texts = []
            for cell in row:
                val = cell.value
                if val is not None:
                    row_texts.append(str(val))
                else:
                    row_texts.append("")

                # л№Ҳ м…Җ к°җм§Җ (мҲҳмӢқ м…Җ м ңмҷё)
                is_empty = val is None or (isinstance(val, str) and val.strip() == "")
                is_formula = isinstance(val, str) and val.startswith("=")

                if is_formula:
                    continue

                if is_empty or include_filled:
                    # мқём ‘ лқјлІЁ м°ҫкё°
                    label = _find_xlsx_label(ws, cell.row, cell.column)
                    if not label and is_empty:
                        # лқјлІЁ м—ҶлҠ” л№Ҳ м…ҖмқҖ л¬ҙмӢң (н…Ңл‘җлҰ¬к°Җ мһҲлҠ” м…Җл§Ң)
                        if not _has_border(cell):
                            continue

                    field_id += 1
                    fields.append({
                        "id": f"xlsx_{field_id}",
                        "sheet": sheet_name,
                        "cell_ref": cell.coordinate,
                        "row": cell.row,
                        "col": cell.column,
                        "label": label or f"({cell.coordinate})",
                        "context": _get_xlsx_context(ws, cell.row, cell.column),
                        "current_value": str(val) if val is not None else "",
                        "is_empty": is_empty,
                        "value_type": "text",
                    })

            if any(t for t in row_texts):
                text_parts.append(" | ".join(row_texts))

    wb.close()
    return fields, "\n".join(text_parts)


def _find_xlsx_label(ws, row: int, col: int) -> str:
    """л№Ҳ м…Җмқҳ мҷјмӘҪ лҳҗлҠ” мң„мӘҪм—җм„ң лқјлІЁ(н…ҚмҠӨнҠё) нғҗмғү."""
    # мҷјмӘҪ м…Җ
    if col > 1:
        left = ws.cell(row=row, column=col - 1).value
        if left and isinstance(left, str) and left.strip():
            return left.strip()
    # мң„мӘҪ м…Җ
    if row > 1:
        above = ws.cell(row=row - 1, column=col).value
        if above and isinstance(above, str) and above.strip():
            return above.strip()
    # мҷјмӘҪ 2м№ё
    if col > 2:
        left2 = ws.cell(row=row, column=col - 2).value
        if left2 and isinstance(left2, str) and left2.strip():
            return left2.strip()
    return ""


def _has_border(cell) -> bool:
    """м…Җм—җ н…Ңл‘җлҰ¬к°Җ мһҲлҠ”м§Җ нҷ•мқё."""
    border = cell.border
    if border is None:
        return False
    for side in [border.left, border.right, border.top, border.bottom]:
        if side and side.style and side.style != "none":
            return True
    return False


def _get_xlsx_context(ws, row: int, col: int) -> str:
    """л№Ҳ м…Җ мЈјліҖмқҳ н…ҚмҠӨнҠёлҘј лӘЁм•„ л§ҘлқҪ м ңкіө."""
    parts = []
    for r in range(max(1, row - 1), min(ws.max_row + 1, row + 2)):
        for c in range(max(1, col - 2), min(ws.max_column + 1, col + 3)):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                parts.append(str(v).strip())
    return " | ".join(parts[:10])


# в”Җв”Җ XLSX fallback (lxml м§Ғм ‘ нҢҢмӢұ) в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ

def _extract_xlsx_raw(path: str, include_filled: bool, context) -> tuple[list, str]:
    """openpyxl мӢӨнҢЁ мӢң lxmlмңјлЎң xlsx XMLмқ„ м§Ғм ‘ нҢҢмӢұн•ҳм—¬ л№Ҳм№ё м¶”м¶ң."""
    import re as _re
    from lxml import etree

    context["log"]("м—‘м…Җ м–‘мӢқ л¶„м„қ мӨ‘ (lxml м§Ғм ‘ нҢҢмӢұ)...")

    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns = {"x": NS}

    fields = []
    text_parts = []
    field_id = 0

    with zipfile.ZipFile(path, "r") as zf:
        # shared strings
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            ss_xml = etree.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in ss_xml.findall(".//x:si", ns):
                shared_strings.append("".join(si.itertext()))

        # мӢңнҠё мқҙлҰ„ лӘ©лЎқ
        wb_xml = etree.fromstring(zf.read("xl/workbook.xml"))
        sheet_names = [s.get("name") for s in wb_xml.findall(".//x:sheets/x:sheet", ns)]

        # мӢңнҠё нҢҢмқј лӘ©лЎқ (relsм—җм„ң)
        rels_xml = etree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_map = {}
        for rel in rels_xml:
            rid_map[rel.get("Id")] = rel.get("Target")

        wb_sheets = wb_xml.findall(".//x:sheets/x:sheet", ns)

        for sheet_idx, sheet_elem in enumerate(wb_sheets):
            sheet_name = sheet_elem.get("name", f"Sheet{sheet_idx+1}")
            rid = sheet_elem.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            target = rid_map.get(rid, f"worksheets/sheet{sheet_idx+1}.xml")
            sheet_path = f"xl/{target}" if not target.startswith("xl/") else target

            if sheet_path not in zf.namelist():
                continue

            sheet_xml = etree.fromstring(zf.read(sheet_path))
            text_parts.append(f"=== мӢңнҠё: {sheet_name} ===")

            # м…Җ лҚ°мқҙн„°лҘј dictлЎң мҲҳм§‘: (row, col) вҶ’ value
            cell_data: dict[tuple[int, int], str] = {}
            max_row = 0
            max_col = 0

            for row_elem in sheet_xml.findall(".//x:sheetData/x:row", ns):
                for c_elem in row_elem.findall("x:c", ns):
                    ref = c_elem.get("r", "")
                    if not ref:
                        continue

                    # м…Җ м°ёмЎ° нҢҢмӢұ (A1 вҶ’ row=1, col=1)
                    m = _re.match(r"([A-Z]+)(\d+)", ref)
                    if not m:
                        continue
                    col_str, row_num = m.group(1), int(m.group(2))
                    col_num = _col_letter_to_num(col_str)
                    max_row = max(max_row, row_num)
                    max_col = max(max_col, col_num)

                    # к°’ м¶”м¶ң
                    t = c_elem.get("t", "")
                    v_elem = c_elem.find("x:v", ns)
                    val = v_elem.text if v_elem is not None else ""

                    if t == "s" and val:
                        idx = int(val)
                        val = shared_strings[idx] if idx < len(shared_strings) else val

                    # мҲҳмӢқ м…Җ (f нғңк·ё)
                    f_elem = c_elem.find("x:f", ns)
                    is_formula = f_elem is not None

                    cell_data[(row_num, col_num)] = val if val else ""

                    if is_formula:
                        continue  # мҲҳмӢқ м…ҖмқҖ л№Ҳм№ё нӣ„ліҙм—җм„ң м ңмҷё

                    is_empty = not val or (isinstance(val, str) and not val.strip())

                    if is_empty or include_filled:
                        # мқём ‘ лқјлІЁ м°ҫкё°
                        label = _find_raw_label(cell_data, row_num, col_num)
                        if not label and is_empty:
                            continue  # лқјлІЁ м—ҶлҠ” л№Ҳ м…Җ л¬ҙмӢң (border нҷ•мқё л¶Ҳк°Җ)

                        field_id += 1
                        context_text = _get_raw_context(cell_data, row_num, col_num)
                        fields.append({
                            "id": f"xlsx_{field_id}",
                            "sheet": sheet_name,
                            "cell_ref": ref,
                            "row": row_num,
                            "col": col_num,
                            "label": label or f"({ref})",
                            "context": context_text,
                            "current_value": val if val else "",
                            "is_empty": is_empty,
                            "value_type": "text",
                        })

            # text_parts кө¬м„ұ
            for r in range(1, max_row + 1):
                row_vals = []
                for c in range(1, max_col + 1):
                    row_vals.append(cell_data.get((r, c), ""))
                if any(v for v in row_vals):
                    text_parts.append(" | ".join(row_vals))

    context["log"](f"lxml нҢҢмӢұ мҷ„лЈҢ: л№Ҳм№ё {len(fields)}к°ң")
    return fields, "\n".join(text_parts)


def _col_letter_to_num(letters: str) -> int:
    """AвҶ’1, BвҶ’2, ..., ZвҶ’26, AAвҶ’27, ..."""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _num_to_col_letter(n: int) -> str:
    """1вҶ’A, 2вҶ’B, ..., 26вҶ’Z, 27вҶ’AA, ..."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _find_raw_label(cell_data: dict, row: int, col: int) -> str:
    """dict кё°л°ҳ мқём ‘ лқјлІЁ м°ҫкё°."""
    # мҷјмӘҪ
    left = cell_data.get((row, col - 1), "")
    if left and left.strip():
        return left.strip()
    # мң„мӘҪ
    above = cell_data.get((row - 1, col), "")
    if above and above.strip():
        return above.strip()
    # мҷјмӘҪ 2м№ё
    left2 = cell_data.get((row, col - 2), "")
    if left2 and left2.strip():
        return left2.strip()
    return ""


def _get_raw_context(cell_data: dict, row: int, col: int) -> str:
    """dict кё°л°ҳ мЈјліҖ л§ҘлқҪ."""
    parts = []
    for r in range(row - 1, row + 2):
        for c in range(col - 2, col + 3):
            v = cell_data.get((r, c), "")
            if v and v.strip():
                parts.append(v.strip())
    return " | ".join(parts[:10])


# в”Җв”Җ HWPX в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ

def _extract_hwpx(path: str, include_filled: bool, context) -> tuple[list, str]:
    """лі‘н•©-мқём§Җ к·ёлҰ¬л“ң м¶”м¶ң (engine/hwpml/hwpx_grid).

    н‘ңлҘј (н–ү,м—ҙ) к·ёлҰ¬л“ң + лі‘н•© л§өмңјлЎң н•ҙм„қн•ҳкі , л№Ҳм№ёл§ҲлӢӨ н–үн—ӨлҚ”Г—м—ҙн—ӨлҚ”лҘј
    мҪ”л“ңлЎң кі„мӮ°н•ңлӢӨ. LLMмқҖ мқҳлҜё л§Өм№ӯл§Ң н•ҳл©ҙ лҗңлӢӨ.
    м…Җ ID нҳ•мӢқ: s{м„№м…ҳ}_t{н‘ң}_r{н–ү}_c{м—ҙ} вҖ” form_fillкіј кіөмң н•ҳлҠ” мЈјмҶҢ мІҙкі„.
    """
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
    from engine.hwpml.hwpx_grid import parse_hwpx, extract_blank_fields

    context["log"]("HWPX м–‘мӢқ л¶„м„қ мӨ‘ (лі‘н•©-мқём§Җ к·ёлҰ¬л“ң)...")

    doc = parse_hwpx(path)
    fields = extract_blank_fields(doc, include_filled=include_filled)
    full_text = doc.render_text(mark_blanks=True)

    # лҲ„лҰ„нӢҖ(нҸј н•„л“ң)мқҖ к·ёлҰ¬л“ңмҷҖ лі„к°ңлЎң ліҙмЎҙ
    from lxml import etree
    field_id = 0
    with zipfile.ZipFile(path, "r") as zf:
        section_files = sorted(
            [n for n in zf.namelist() if "section" in n.lower() and n.endswith(".xml")]
        )
        for sec_file in section_files:
            try:
                root = etree.fromstring(zf.read(sec_file))
            except etree.XMLSyntaxError:
                continue
            for elem in root.iter():
                tag = etree.QName(elem.tag).localname if "}" in str(elem.tag) else str(elem.tag)
                if tag in ("fieldBegin", "FIELDBEGIN"):
                    name = elem.get("name", "") or elem.get("Name", "")
                    field_id += 1
                    fields.append({
                        "id": f"hwpx_field_{field_id}",
                        "label": name or f"н•„л“ң{field_id}",
                        "context": "",
                        "field_name": name,
                        "xpath": sec_file,
                        "current_value": "",
                        "is_empty": True,
                        "value_type": "field",
                    })

    context["log"](f"н‘ң {len(doc.tables)}к°ң, к·ёлҰ¬л“ң л№Ҳм№ё {sum(1 for f in fields if f.get('value_type') == 'text')}к°ң")
    return fields, full_text


# в”Җв”Җ HWP (COM) в”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җв”Җ

def _extract_hwp(path: str, include_filled: bool, context) -> tuple[list, str]:
    """HWP л°”мқҙл„ҲлҰ¬ вҖ” olefileлЎң н…ҚмҠӨнҠё м¶”м¶ң (COM л¶Ҳн•„мҡ”)."""
    context["log"]("HWP н…ҚмҠӨнҠё л¶„м„қ мӨ‘...")

    # olefileлЎң н…ҚмҠӨнҠё м¶”м¶ң
    full_text = ""
    try:
        import olefile
        ole = olefile.OleFileIO(path)
        if ole.exists("PrvText"):
            data = ole.openstream("PrvText").read()
            full_text = data.decode("utf-16-le", errors="ignore")
        else:
            texts = []
            for stream in ole.listdir():
                name = "/".join(stream)
                if "BodyText" in name:
                    raw = ole.openstream(stream).read()
                    text = raw.decode("utf-16-le", errors="ignore")
                    text = "".join(c for c in text if c.isprintable() or c in "\n\r\t")
                    if text.strip():
                        texts.append(text.strip())
            full_text = "\n\n".join(texts)
        ole.close()
    except Exception as e:
        context["log"](f"olefile м¶”м¶ң мӢӨнҢЁ: {e}")

    if not full_text:
        try:
            from hwp5.hwp5txt import extract_text
            import io
            buf = io.StringIO()
            extract_text(path, buf)
            full_text = buf.getvalue()
        except Exception:
            pass

    # м¶”м¶ң мҷ„м „ мӢӨнҢЁ мӢң л¬ҙмқҢ м„ұкіө(л№Ҳ м–‘мӢқ) лҢҖмӢ  лӘ…нҷ•нһҲ мӢӨнҢЁмӢңнӮЁлӢӨ
    if not full_text or not full_text.strip():
        raise RuntimeError(
            f"HWP н…ҚмҠӨнҠё м¶”м¶ң мӢӨнҢЁ: {os.path.basename(path)} вҖ” "
            f"olefile/hwp5 лҜём„Өм№ҳмқҙкұ°лӮҳ м§Җмӣҗн•ҳм§Җ м•ҠлҠ” нҳ•мӢқмһ…лӢҲлӢӨ. "
            f".hwpxлЎң м ҖмһҘ нӣ„ мӮ¬мҡ©н•ҳкұ°лӮҳ 'pip install olefile'мқ„ м„Өм№ҳн•ҳм„ёмҡ”."
        )

    # н…ҚмҠӨнҠё кё°л°ҳ л№Ҳм№ё м¶”м¶ң вҖ” л№Ҳ кҙ„нҳё, л°‘мӨ„, л№Ҳ м№ё нҢЁн„ҙ
    import re
    fields = []
    field_id = 0
    lines = full_text.split("\n")
    prev_line = ""

    for line_num, line in enumerate(lines, 1):
        stripped = line.strip()

        # л№Ҳ кҙ„нҳё нҢЁн„ҙ: (     ), [     ], гҖҢ    гҖҚ
        for m in re.finditer(r'[\(\[гҖҢ]\s{2,}[\)\]гҖҚ]', stripped):
            field_id += 1
            label = prev_line.strip()[-20:] if prev_line.strip() else f"н–ү{line_num}"
            fields.append({
                "id": f"hwp_{field_id}",
                "label": label,
                "context": stripped[:50],
                "line": line_num,
                "current_value": "",
                "is_empty": True,
                "value_type": "text",
            })

        # л°‘мӨ„ нҢЁн„ҙ: _____
        for m in re.finditer(r'_{3,}', stripped):
            field_id += 1
            # л°‘мӨ„ м•һмқҳ н…ҚмҠӨнҠёлҘј лқјлІЁлЎң
            before = stripped[:m.start()].strip()
            label = before[-20:] if before else f"н–ү{line_num}"
            fields.append({
                "id": f"hwp_{field_id}",
                "label": label,
                "context": stripped[:50],
                "line": line_num,
                "current_value": "",
                "is_empty": True,
                "value_type": "text",
            })

        if stripped:
            prev_line = stripped

    context["log"](f"н…ҚмҠӨнҠё {len(lines)}мӨ„, л№Ҳм№ё нҢЁн„ҙ {len(fields)}к°ң")
    return fields, full_text
