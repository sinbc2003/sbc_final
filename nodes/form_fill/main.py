"""
양식 값 주입 노드.

XLSX: openpyxl로 셀 값 쓰기 (서식/수식/매크로 보존)
HWPX: ZIP→XML 내 텍스트 교체
HWP: 한/글 COM API (누름틀 채우기)
"""

import json
import os
import shutil
import zipfile
from pathlib import Path


def execute(inputs: dict, params: dict, context: dict) -> dict:
    form_path = inputs["양식파일"]
    fill_json = inputs["채울내용"]

    if not Path(form_path).exists():
        raise FileNotFoundError(f"양식 파일 없음: {form_path}")

    # JSON 파싱
    try:
        fill_data = json.loads(fill_json)
    except json.JSONDecodeError:
        raise ValueError("채울내용이 올바른 JSON이 아닙니다.")

    if not isinstance(fill_data, dict):
        raise ValueError("채울내용은 {빈칸ID: 값} 형태의 JSON 객체여야 합니다.")

    ext = Path(form_path).suffix.lower()

    # 구형 .xls(OLE 바이너리)는 openpyxl 미지원 → 혼란스러운 BadZipFile 대신 명확히 안내
    if ext == ".xls":
        raise ValueError(
            "구형 .xls 형식은 지원하지 않습니다. 엑셀에서 .xlsx로 저장한 뒤 사용하세요."
        )

    output_name = params.get("output_name", "완성")
    output_ext = ext if ext in (".xlsx", ".hwpx") else ".hwpx"
    output_path = os.path.join(context["temp_dir"], f"{output_name}{output_ext}")

    context["progress"](0.1)

    if ext == ".xlsx":
        _fill_xlsx(form_path, fill_data, output_path, context)
    elif ext == ".hwpx":
        _fill_hwpx(form_path, fill_data, output_path, context)
    elif ext == ".hwp":
        # HWP는 COM 필요 — DLL 미설치 시 텍스트 결과만 반환
        try:
            _fill_hwp(form_path, fill_data, output_path, context)
        except Exception as e:
            context["log"](f"HWP COM 사용 불가: {e}")
            context["log"]("채울 내용을 텍스트로 반환합니다 (수동으로 한/글에 붙여넣기)")
            # 텍스트 파일로 결과 저장
            txt_path = output_path.rsplit(".", 1)[0] + "_채울내용.txt"
            with open(txt_path, "w", encoding="utf-8") as f:
                for k, v in fill_data.items():
                    if str(v).strip():
                        f.write(f"{k}: {v}\n")
            context["log"](f"텍스트 저장: {txt_path}")
            return {"파일": txt_path}
    else:
        raise ValueError(f"지원하지 않는 형식: {ext}")

    context["progress"](1.0)
    context["log"](f"양식 작성 완료: {len(fill_data)}개 항목")

    return {"파일": output_path}


# ── XLSX ──────────────────────────────────────────────

def _fill_xlsx(form_path: str, fill_data: dict, output_path: str, context):
    import openpyxl

    context["log"]("엑셀 양식에 값 주입 중...")

    # 원본 복사 후 편집 (매크로 보존)
    shutil.copy2(form_path, output_path)

    # openpyxl 스타일 파싱 실패 시 lxml 직접 편집 fallback
    keep_vba = form_path.lower().endswith(".xlsm")
    try:
        wb = openpyxl.load_workbook(output_path, keep_vba=keep_vba)
    except (IndexError, Exception) as e:
        context["log"](f"openpyxl 로드 실패, lxml 직접 편집으로 전환: {e}")
        _fill_xlsx_raw(form_path, fill_data, output_path, context)
        return

    filled = 0
    for key, value in fill_data.items():
        # key 형태: "xlsx_1", "B3", "시트1!B3", "시트1_B3"
        sheet_name = None
        cell_ref = None

        if "!" in key:
            # "시트1!B3"
            sheet_name, cell_ref = key.split("!", 1)
        elif key.startswith("xlsx_"):
            # form_extract의 ID — fill_data에서 cell_ref를 찾아야 함
            # 이 경우 fill_data의 value가 dict일 수 있음
            if isinstance(value, dict):
                sheet_name = value.get("sheet")
                cell_ref = value.get("cell_ref")
                value = value.get("value", "")
            else:
                context["log"](f"경고: '{key}' — cell_ref 없음, 건너뜀")
                continue
        else:
            # 직접 셀 참조 ("B3")
            cell_ref = key

        if not cell_ref:
            continue

        # 시트 찾기
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        try:
            cell = ws[cell_ref]
            # 병합 셀이면 좌상단 셀에 쓰기
            if hasattr(cell, 'coordinate') and cell.__class__.__name__ == 'MergedCell':
                for merged_range in ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        break
            cell.value = value
            filled += 1
        except Exception as e:
            context["log"](f"경고: {cell_ref} 쓰기 실패 - {e}")

    wb.save(output_path)
    wb.close()
    context["log"](f"엑셀 {filled}개 셀 주입 완료")


def _fill_xlsx_raw(form_path: str, fill_data: dict, output_path: str, context):
    """openpyxl 실패 시 lxml으로 xlsx XML을 직접 편집하여 값 주입."""
    import re as _re
    from lxml import etree

    context["log"]("엑셀 양식에 값 주입 중 (lxml 직접 편집)...")

    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns = {"x": NS}

    # shared strings 로드
    shared_strings: list[str] = []
    with zipfile.ZipFile(form_path, "r") as zf:
        if "xl/sharedStrings.xml" in zf.namelist():
            ss_xml = etree.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in ss_xml.findall(".//x:si", ns):
                shared_strings.append("".join(si.itertext()))

        # 시트 이름 매핑
        wb_xml = etree.fromstring(zf.read("xl/workbook.xml"))
        sheet_elems = wb_xml.findall(".//x:sheets/x:sheet", ns)
        sheet_name_to_idx = {s.get("name"): i for i, s in enumerate(sheet_elems)}

        rels_xml = etree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_map = {}
        for rel in rels_xml:
            rid_map[rel.get("Id")] = rel.get("Target")

        # 시트별 fill_data 분류
        # key 형태: "B3", "시트1!B3", "xlsx_1" (cell_ref)
        sheet_fills: dict[str, dict[str, str]] = {}  # sheet_path → {cell_ref → value}

        for key, value in fill_data.items():
            sheet_name = None
            cell_ref = None

            if "!" in key:
                sheet_name, cell_ref = key.split("!", 1)
            elif key.startswith("xlsx_"):
                if isinstance(value, dict):
                    sheet_name = value.get("sheet")
                    cell_ref = value.get("cell_ref")
                    value = value.get("value", "")
                else:
                    continue
            else:
                cell_ref = key

            if not cell_ref:
                continue

            val = str(value).strip() if not isinstance(value, dict) else str(value)
            if not val:
                continue

            # 시트 경로 결정
            if sheet_name and sheet_name in sheet_name_to_idx:
                idx = sheet_name_to_idx[sheet_name]
            else:
                idx = 0  # 기본: 첫 번째 시트

            if idx < len(sheet_elems):
                rid = sheet_elems[idx].get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                target = rid_map.get(rid, f"worksheets/sheet{idx+1}.xml")
                sp = f"xl/{target}" if not target.startswith("xl/") else target
            else:
                sp = f"xl/worksheets/sheet{idx+1}.xml"

            sheet_fills.setdefault(sp, {})[cell_ref] = val

        # ZIP 복사하면서 시트 XML 수정
        filled = 0
        with zipfile.ZipFile(form_path, "r") as zf_in:
            with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf_out:
                for item in zf_in.infolist():
                    data = zf_in.read(item.filename)

                    if item.filename in sheet_fills:
                        sheet_xml = etree.fromstring(data)
                        fills_for_sheet = sheet_fills[item.filename]

                        # 기존 셀 수정
                        for row_elem in sheet_xml.findall(".//x:sheetData/x:row", ns):
                            for c_elem in row_elem.findall("x:c", ns):
                                ref = c_elem.get("r", "")
                                if ref in fills_for_sheet:
                                    new_val = fills_for_sheet.pop(ref)
                                    # 인라인 문자열로 설정 (t="inlineStr")
                                    c_elem.set("t", "inlineStr")
                                    # 기존 v 제거
                                    for v in c_elem.findall("x:v", ns):
                                        c_elem.remove(v)
                                    # 기존 is 제거
                                    for is_elem in c_elem.findall("x:is", ns):
                                        c_elem.remove(is_elem)
                                    # 새 inlineStr
                                    is_new = etree.SubElement(c_elem, f"{{{NS}}}is")
                                    t_new = etree.SubElement(is_new, f"{{{NS}}}t")
                                    t_new.text = new_val
                                    filled += 1

                        # 남은 셀 (기존에 없는 셀 → 새로 추가)
                        if fills_for_sheet:
                            sheet_data = sheet_xml.find(".//x:sheetData", ns)
                            if sheet_data is not None:
                                for ref, val in fills_for_sheet.items():
                                    m = _re.match(r"([A-Z]+)(\d+)", ref)
                                    if not m:
                                        continue
                                    row_num = int(m.group(2))
                                    # 해당 행 찾기 또는 생성
                                    row_elem = None
                                    for r in sheet_data.findall("x:row", ns):
                                        if r.get("r") == str(row_num):
                                            row_elem = r
                                            break
                                    if row_elem is None:
                                        row_elem = etree.SubElement(sheet_data, f"{{{NS}}}row")
                                        row_elem.set("r", str(row_num))

                                    c_new = etree.SubElement(row_elem, f"{{{NS}}}c")
                                    c_new.set("r", ref)
                                    c_new.set("t", "inlineStr")
                                    is_new = etree.SubElement(c_new, f"{{{NS}}}is")
                                    t_new = etree.SubElement(is_new, f"{{{NS}}}t")
                                    t_new.text = val
                                    filled += 1

                        data = etree.tostring(sheet_xml, xml_declaration=True, encoding="UTF-8", standalone=True)

                    zf_out.writestr(item, data)

    context["log"](f"엑셀 {filled}개 셀 주입 완료 (lxml)")


# ── HWPX ─────────────────────────────────────────────

def _fill_hwpx(form_path: str, fill_data: dict, output_path: str, context):
    from lxml import etree

    context["log"]("HWPX 양식에 값 주입 중...")

    # ── 그리드 좌표 ID 경로 (form_extract의 병합-인지 그리드와 짝) ──
    # fill_data 키가 "s0_t1_r4_c2" 형식이면 셀 주소로 정확히 주입
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
    from engine.hwpml.hwpx_grid import (
        ID_RE, fill_hwpx_cells, parse_hwpx, relocate_below_markers,
    )

    grid_map = {}
    legacy_map = {}
    for key, value in fill_data.items():
        if isinstance(value, dict):
            value = value.get("value", "")
        if ID_RE.match(str(key).strip()):
            grid_map[str(key).strip()] = str(value)
        else:
            legacy_map[key] = str(value)

    filled = 0
    if grid_map:
        # "이하빈칸" 마커를 채운 행 아래로 자동 이동 (공문서 관례)
        try:
            doc = parse_hwpx(form_path)
            for k, v in relocate_below_markers(doc, grid_map, log=context["log"]).items():
                grid_map.setdefault(k, v)
        except Exception as e:
            context["log"](f"'이하빈칸' 마커 이동 건너뜀: {e}")
        filled += fill_hwpx_cells(form_path, output_path, grid_map, log=context["log"])
        # 이후 legacy 처리는 방금 쓴 output을 입력으로
        work_input = output_path
    else:
        work_input = form_path

    # ── legacy 경로: 누름틀 필드명 / 라벨 휴리스틱 ──
    if legacy_map or not grid_map:
        shutil.copy2(work_input, output_path + ".tmp")
        with zipfile.ZipFile(output_path + ".tmp", "r") as zf_in:
            with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf_out:
                for item in zf_in.infolist():
                    data = zf_in.read(item.filename)
                    if legacy_map and "section" in item.filename.lower() and item.filename.endswith(".xml"):
                        try:
                            root = etree.fromstring(data)
                            count = _fill_hwpx_section(root, legacy_map, legacy_map)
                            filled += count
                            data = etree.tostring(root, xml_declaration=True, encoding="UTF-8")
                        except etree.XMLSyntaxError:
                            pass
                    zf_out.writestr(item, data)
        os.remove(output_path + ".tmp")

    context["log"](f"HWPX {filled}개 필드 주입 완료")


def _fill_hwpx_section(root, field_map: dict, label_map: dict) -> int:
    """section XML에서 누름틀 필드와 빈 텍스트를 채운다."""
    from lxml import etree  # 모듈 레벨 함수라 _fill_hwpx의 지역 import를 못 봄

    filled = 0
    prev_text = ""

    for elem in root.iter():
        tag = etree.QName(elem.tag).localname if "}" in elem.tag else elem.tag

        # 누름틀 필드 값 주입
        if tag in ("fieldBegin", "FIELDBEGIN"):
            name = elem.get("name", "") or elem.get("Name", "")
            # 다음 텍스트 노드에 값을 넣어야 하므로 마킹
            if name:
                # field_map에서 field_name이나 ID로 검색
                for key, val in field_map.items():
                    if name in key or key in name:
                        # 다음 <hp:t> 노드를 찾아서 값 설정
                        next_t = elem.getnext()
                        while next_t is not None:
                            next_tag = etree.QName(next_t.tag).localname if "}" in next_t.tag else next_t.tag
                            if next_tag == "t":
                                next_t.text = val
                                filled += 1
                                break
                            next_t = next_t.getnext()
                        break

        # 빈 텍스트 매칭 — 라벨 기반
        if tag == "t":
            text = elem.text or ""
            if text.strip() == "" and prev_text.strip():
                # 이전 텍스트(라벨)로 매칭
                label = prev_text.strip()
                for key, val in label_map.items():
                    if key in label or label in key:
                        elem.text = val
                        filled += 1
                        break
            if text.strip():
                prev_text = text

    return filled


# ── HWP (COM) ────────────────────────────────────────

def _fill_hwp(form_path: str, fill_data: dict, output_path: str, context):
    """HWP 바이너리 — 한/글 COM API로 누름틀 값 채우기."""
    import platform

    if platform.system() != "Windows":
        raise RuntimeError("HWP 파일은 Windows + 한/글 설치 환경에서만 처리 가능합니다.")

    context["log"]("한/글 COM API로 HWP 양식 작성 중...")

    try:
        import pythoncom
        pythoncom.CoInitialize()
        import win32com.client
    except ImportError:
        raise RuntimeError("pywin32가 필요합니다: pip install pywin32")

    hwp = None
    try:
        hwp = win32com.client.gencache.EnsureDispatch("HWPFrame.HwpObject")
        # ModuleType은 반드시 "FilePathCheckDLL" (예제 DLL 이름이 아님).
        # 등록 실패 시 Open/SaveAs에서 파일접근 승인 대화상자가 떠 무인 실행이 멈춤.
        if not hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule"):
            context["log"](
                "[WARN] 보안모듈(FilePathCheckerModule) 등록 실패 — "
                "한/글 파일접근 승인 대화상자가 뜰 수 있습니다. "
                "pyhwpx 설치로 레지스트리 등록이 필요합니다."
            )
        hwp.Open(os.path.abspath(form_path), "HWP", "forceopen:true")

        filled = 0

        # 누름틀에 값 채우기
        for key, value in fill_data.items():
            if isinstance(value, dict):
                field_name = value.get("field_name", key)
                value = value.get("value", "")
            else:
                field_name = key

            try:
                hwp.PutFieldText(field_name, str(value))
                filled += 1
            except Exception as e:
                context["log"](f"경고: 필드 '{field_name}' 채우기 실패 — {e}")

        # HWPX로 저장 (HWP 원본 보존하면서 HWPX 출력)
        if output_path.endswith(".hwpx"):
            hwp.SaveAs(os.path.abspath(output_path), "HWPX")
        else:
            hwp.SaveAs(os.path.abspath(output_path), "HWP")

        context["log"](f"HWP {filled}개 누름틀 주입 완료")

    finally:
        if hwp:
            try:
                hwp.Clear(1)
                hwp.Quit()
            except Exception:
                pass
